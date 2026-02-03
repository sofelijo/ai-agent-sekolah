from __future__ import annotations

import calendar
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, TYPE_CHECKING

from utils import INDONESIAN_MONTH_NAMES
import re

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError:  # pragma: no cover - handled by caller
    load_workbook = None
    from_excel = None

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
else:
    try:  # pragma: no cover - only for runtime
        from openpyxl.cell.cell import MergedCell
    except Exception:  # pragma: no cover
        MergedCell = None  # type: ignore


SEMESTER_2_2025_2026_MONTHS: Tuple[Tuple[int, int], ...] = (
    (2026, 1),
    (2026, 2),
    (2026, 3),
    (2026, 4),
    (2026, 5),
    (2026, 6),
)

DATE_START_COLUMN = 4  # Column D
DATE_COLUMNS = 31


def _format_academic_year(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    if "/" in cleaned:
        parts = [part.strip() for part in cleaned.split("/") if part.strip()]
        if len(parts) == 2:
            return f"{parts[0]} - {parts[1]}"
    if "-" in cleaned:
        parts = [part.strip() for part in cleaned.split("-") if part.strip()]
        if len(parts) == 2 and all(part.isdigit() for part in parts):
            return f"{parts[0]} - {parts[1]}"
    return cleaned


def _format_class_label(class_name: Optional[str]) -> Optional[str]:
    if not class_name:
        return None
    cleaned = class_name.strip()
    if not cleaned:
        return None
    upper = cleaned.upper()
    if upper.startswith("KELAS"):
        return cleaned
    return f"KELAS {cleaned}"


def _compose_teacher_name(full_name: Optional[str], prefix: Optional[str], suffix: Optional[str]) -> Optional[str]:
    parts = []
    if prefix:
        prefix_clean = prefix.strip()
        if prefix_clean:
            parts.append(prefix_clean)
    if full_name:
        name_clean = full_name.strip()
        if name_clean:
            parts.append(name_clean)
    if suffix:
        suffix_clean = suffix.strip()
        if suffix_clean:
            parts.append(suffix_clean)
    return " ".join(parts) if parts else None


def _find_cell(ws: "Worksheet", predicate, *, max_rows: Optional[int] = None):
    for row in ws.iter_rows(min_row=1, max_row=max_rows):
        for cell in row:
            if predicate(cell.value):
                return cell
    return None


def _find_header_row(ws: "Worksheet") -> int:
    cell = _find_cell(
        ws,
        lambda value: isinstance(value, str) and value.strip().upper() == "NO",
        max_rows=25,
    )
    return cell.row if cell else 6


def _find_date_start_col(ws: "Worksheet", header_row: int) -> int:
    cell = _find_cell(
        ws,
        lambda value: isinstance(value, str) and value.strip().upper() == "TANGGAL",
        max_rows=header_row + 1,
    )
    return int(cell.column) if cell else DATE_START_COLUMN


def _find_bulan_cell(ws: "Worksheet"):
    return _find_cell(
        ws,
        lambda value: isinstance(value, str) and value.strip().lower().startswith("bulan"),
        max_rows=10,
    )


def _parse_bulan_label(value: Optional[str]) -> Optional[Tuple[int, int]]:
    if not value or not isinstance(value, str):
        return None
    text = value.lower()
    year_match = re.search(r"(\\d{4})", text)
    if not year_match:
        return None
    year = int(year_match.group(1))
    month = None
    for num, label in INDONESIAN_MONTH_NAMES.items():
        if label.lower() in text:
            month = num
            break
    if not month:
        return None
    return year, month


def _find_nip_row(ws: "Worksheet") -> Optional[int]:
    cell = _find_cell(
        ws,
        lambda value: isinstance(value, str) and value.strip().upper().startswith("NIP"),
    )
    return cell.row if cell else None


def _find_ket_row(ws: "Worksheet") -> Optional[int]:
    cell = _find_cell(
        ws,
        lambda value: isinstance(value, str) and value.strip().lower() == "ket :",
    )
    return cell.row if cell else None


def _find_max_col(ws: "Worksheet", header_row: int) -> int:
    max_col = 1
    for cell in ws[header_row]:
        if cell.value not in (None, ""):
            max_col = max(max_col, int(cell.column))
    return max_col


def _copy_row_style(ws: "Worksheet", src_row: int, dest_row: int, max_col: int) -> None:
    ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dest = ws.cell(row=dest_row, column=col)
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = src.number_format
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)


def _clone_style(cell) -> Dict[str, Any]:
    return {
        "font": copy(cell.font),
        "border": copy(cell.border),
        "fill": copy(cell.fill),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
        "alignment": copy(cell.alignment),
    }


def _apply_style(cell, style: Dict[str, Any]) -> None:
    cell.font = style["font"]
    cell.border = style["border"]
    cell.fill = style["fill"]
    cell.number_format = style["number_format"]
    cell.protection = style["protection"]
    cell.alignment = style["alignment"]


def _clear_range(ws: "Worksheet", start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None


def _coerce_date(value, wb) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and from_excel is not None and wb is not None:
        try:
            return from_excel(value, wb.epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except Exception:
            return None
    return None


def _extract_weekday_weekend_styles(
    ws: "Worksheet",
    wb: "Workbook",
    *,
    date_row: int,
    date_start_col: int,
    reference_row: int,
    template_month_year: Optional[Tuple[int, int]] = None,
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    weekday_style = None
    weekend_style = None

    if template_month_year:
        year, month = template_month_year
        days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, days_in_month + 1):
            col = date_start_col + day - 1
            ref_cell = ws.cell(row=reference_row, column=col)
            if MergedCell is not None and isinstance(ref_cell, MergedCell):
                continue
            style = _clone_style(ref_cell)
            if date(year, month, day).weekday() >= 5:
                if weekend_style is None:
                    weekend_style = style
            else:
                if weekday_style is None:
                    weekday_style = style
            if weekday_style and weekend_style:
                break
    else:
        for offset in range(DATE_COLUMNS):
            col = date_start_col + offset
            date_cell = ws.cell(row=date_row, column=col)
            dt = _coerce_date(date_cell.value, wb)
            if not dt:
                continue
            ref_cell = ws.cell(row=reference_row, column=col)
            if MergedCell is not None and isinstance(ref_cell, MergedCell):
                continue
            style = _clone_style(ref_cell)
            if dt.weekday() >= 5:
                if weekend_style is None:
                    weekend_style = style
            else:
                if weekday_style is None:
                    weekday_style = style
            if weekday_style and weekend_style:
                break

    return weekday_style, weekend_style


def _apply_weekend_styles(
    ws: "Worksheet",
    *,
    year: int,
    month: int,
    date_row: int,
    date_start_col: int,
    student_start_row: int,
    student_end_row: int,
    date_weekday_style: Optional[Dict[str, Any]],
    date_weekend_style: Optional[Dict[str, Any]],
    student_weekday_style: Optional[Dict[str, Any]],
    student_weekend_style: Optional[Dict[str, Any]],
) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        col = date_start_col + day - 1
        is_weekend = date(year, month, day).weekday() >= 5
        header_style = date_weekday_style if is_weekend else date_weekend_style
        if header_style:
            cell = ws.cell(row=date_row, column=col)
            if MergedCell is None or not isinstance(cell, MergedCell):
                _apply_style(cell, header_style)
        row_style = student_weekday_style if is_weekend else student_weekend_style
        if row_style:
            for row in range(student_start_row, student_end_row + 1):
                cell = ws.cell(row=row, column=col)
                if MergedCell is not None and isinstance(cell, MergedCell):
                    continue
                _apply_style(cell, row_style)


def _apply_static_context(
    ws: "Worksheet",
    *,
    students: List[Dict[str, Any]],
    school_name: Optional[str],
    academic_year: Optional[str],
    class_label: Optional[str],
    teacher_name: Optional[str],
    teacher_nip: Optional[str],
    headmaster_name: Optional[str],
    headmaster_nip: Optional[str],
) -> Tuple[int, int, int, int]:
    if school_name:
        ws["A2"] = school_name

    academic_label = _format_academic_year(academic_year)
    if academic_label:
        ws["A3"] = f"TAHUN PELAJARAN {academic_label}"

    class_value = _format_class_label(class_label)
    if class_value:
        ws["B5"] = class_value

    header_row = _find_header_row(ws)
    date_row = header_row + 1
    max_col = _find_max_col(ws, date_row)

    student_start_row = header_row + 2
    nip_row = _find_nip_row(ws)
    ket_row = _find_ket_row(ws)
    jumlah_row = _find_row_by_label(ws, "JUMLAH SISWA HADIR", start_row=student_start_row)
    recap_title_row = _find_row_by_label(ws, "REKAPITULASI KETIDAKHADIRAN TIAP HARI", start_row=student_start_row)
    reserved_candidates = [row for row in (jumlah_row, recap_title_row, nip_row, ket_row) if row]
    reserved_row = min(reserved_candidates) if reserved_candidates else None

    if reserved_row:
        student_end_row = reserved_row - 1
        capacity = student_end_row - student_start_row + 1
        if len(students) > capacity:
            extra_rows = len(students) - capacity
            ws.insert_rows(reserved_row, extra_rows)
            for row in range(reserved_row, reserved_row + extra_rows):
                _copy_row_style(ws, student_start_row, row, max_col)
            if nip_row:
                nip_row += extra_rows
            if ket_row:
                ket_row += extra_rows
            if jumlah_row:
                jumlah_row += extra_rows
            if recap_title_row:
                recap_title_row += extra_rows
            reserved_row += extra_rows
            student_end_row = reserved_row - 1
    else:
        student_end_row = student_start_row + max(len(students), 40) - 1

    _clear_range(ws, student_start_row, student_end_row, 1, max_col)

    for idx, student in enumerate(students, start=1):
        row = student_start_row + idx - 1
        ws.cell(row=row, column=1).value = student.get("no", idx)
        ws.cell(row=row, column=2).value = student.get("name")
        ws.cell(row=row, column=3).value = student.get("gender")

    # Signature blocks
    kepsek_row = _find_row_by_label(ws, "KEPALA SEKOLAH", start_row=1)
    guru_row = None
    for row in ws.iter_rows(min_row=1, max_row=80):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip().upper().startswith("GURU KELAS"):
                guru_row = int(cell.row)
                guru_col = int(cell.column)
                break
        if guru_row:
            break

    if kepsek_row:
        kepsek_col = None
        for cell in ws[kepsek_row]:
            if isinstance(cell.value, str) and cell.value.strip().upper() == "KEPALA SEKOLAH":
                kepsek_col = int(cell.column)
                break
        if kepsek_col:
            if headmaster_name:
                ws.cell(row=kepsek_row + 5, column=kepsek_col).value = headmaster_name
            nip_value = f"NIP. {headmaster_nip}" if headmaster_nip else "NIP."
            ws.cell(row=kepsek_row + 6, column=kepsek_col).value = nip_value

    if guru_row and guru_col:
        if class_label:
            ws.cell(row=guru_row, column=guru_col).value = f"Guru Kelas {class_label}"
        if teacher_name:
            ws.cell(row=guru_row + 5, column=guru_col).value = teacher_name
        nip_value = f"NIP. {teacher_nip}" if teacher_nip else "NIP."
        ws.cell(row=guru_row + 6, column=guru_col).value = nip_value

    return header_row, date_row, student_start_row, student_end_row


def _apply_month_context(
    ws: "Worksheet",
    *,
    year: int,
    month: int,
    header_row: int,
    date_row: int,
) -> None:
    date_start_col = _find_date_start_col(ws, header_row)
    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, DATE_COLUMNS + 1):
        cell = ws.cell(row=date_row, column=date_start_col + day - 1)
        if day <= days_in_month:
            cell.value = date(year, month, day)
        else:
            cell.value = None

    bulan_cell = _find_bulan_cell(ws)
    if bulan_cell:
        month_label = INDONESIAN_MONTH_NAMES.get(month, str(month))
        bulan_cell.value = f"Bulan : {month_label} {year}"


def _find_recap_columns(ws: "Worksheet", date_row: int) -> Dict[str, int]:
    recap_cols: Dict[str, int] = {}
    for cell in ws[date_row]:
        value = cell.value
        if not isinstance(value, str):
            continue
        label = value.strip().upper()
        if label in {"HADIR", "SAKIT", "IZIN", "ALFA", "%"}:
            recap_cols[label] = int(cell.column)
    return recap_cols


def _find_row_by_label(ws: "Worksheet", label: str, *, start_row: int = 1) -> Optional[int]:
    label_norm = label.strip().upper()
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip().upper() == label_norm:
                return int(cell.row)
    return None


def _apply_attendance_data(
    ws: "Worksheet",
    *,
    students: List[Dict[str, Any]],
    month_entries: Dict[Tuple[int, int], str],
    header_row: int,
    student_start_row: int,
    date_start_col: int,
    year: int,
    month: int,
) -> None:
    days_in_month = calendar.monthrange(year, month)[1]
    recap_cols = _find_recap_columns(ws, header_row + 1)
    symbol_map = {"masuk": "✓", "sakit": "S", "izin": "I", "alpa": "A"}

    # Effective days: days with any entry for the class in this month.
    effective_days = len({day for (_, day), _ in month_entries.items()})
    daily_counts = {day: {"masuk": 0, "sakit": 0, "izin": 0, "alpa": 0} for day in range(1, days_in_month + 1)}

    for idx, student in enumerate(students):
        row = student_start_row + idx
        student_id = int(student.get("id"))
        totals = {"masuk": 0, "sakit": 0, "izin": 0, "alpa": 0}
        has_any = False
        for day in range(1, days_in_month + 1):
            status = month_entries.get((student_id, day))
            if not status:
                continue
            normalized = status.strip().lower()
            symbol = symbol_map.get(normalized)
            if symbol:
                cell = ws.cell(row=row, column=date_start_col + day - 1)
                if MergedCell is None or not isinstance(cell, MergedCell):
                    cell.value = symbol
            if normalized in totals:
                totals[normalized] += 1
                daily_counts[day][normalized] += 1
                has_any = True

        hadir = totals["masuk"]
        if not has_any:
            continue

        present_percent = round((hadir / effective_days) * 100) if effective_days else 0

        if "HADIR" in recap_cols:
            ws.cell(row=row, column=recap_cols["HADIR"]).value = hadir
        if "SAKIT" in recap_cols:
            ws.cell(row=row, column=recap_cols["SAKIT"]).value = totals["sakit"]
        if "IZIN" in recap_cols:
            ws.cell(row=row, column=recap_cols["IZIN"]).value = totals["izin"]
        if "ALFA" in recap_cols:
            ws.cell(row=row, column=recap_cols["ALFA"]).value = totals["alpa"]
        if "%" in recap_cols:
            cell = ws.cell(row=row, column=recap_cols["%"])
            cell.value = present_percent / 100

    # Fill summary: jumlah siswa hadir (per day) and recap ketidakhadiran
    jumlah_row = _find_row_by_label(ws, "JUMLAH SISWA HADIR", start_row=header_row + 1)
    if jumlah_row:
        for day in range(1, days_in_month + 1):
            ws.cell(row=jumlah_row, column=date_start_col + day - 1).value = daily_counts[day]["masuk"]
        total_hadir = sum(daily_counts[day]["masuk"] for day in range(1, days_in_month + 1))
        total_col = date_start_col + DATE_COLUMNS
        ws.cell(row=jumlah_row, column=total_col).value = total_hadir

    recap_title_row = _find_row_by_label(ws, "REKAPITULASI KETIDAKHADIRAN TIAP HARI", start_row=header_row + 1)
    recap_start = recap_title_row + 1 if recap_title_row else header_row + 1
    sakit_row = _find_row_by_label(ws, "SAKIT", start_row=recap_start)
    izin_row = _find_row_by_label(ws, "IZIN", start_row=recap_start)
    alfa_row = _find_row_by_label(ws, "ALFA", start_row=recap_start)

    def fill_recap(row_num: Optional[int], key: str) -> None:
        if not row_num:
            return
        for day in range(1, days_in_month + 1):
            ws.cell(row=row_num, column=date_start_col + day - 1).value = daily_counts[day][key]
        total_col = date_start_col + DATE_COLUMNS
        ws.cell(row=row_num, column=total_col).value = sum(daily_counts[day][key] for day in range(1, days_in_month + 1))

    fill_recap(sakit_row, "sakit")
    fill_recap(izin_row, "izin")
    fill_recap(alfa_row, "alpa")


def _clear_attendance_grid(
    ws: "Worksheet",
    *,
    student_start_row: int,
    student_end_row: int,
    date_start_col: int,
    recap_cols: Dict[str, int],
) -> None:
    end_col = date_start_col + DATE_COLUMNS - 1
    for row in ws.iter_rows(
        min_row=student_start_row,
        max_row=student_end_row,
        min_col=date_start_col,
        max_col=end_col,
    ):
        for cell in row:
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None
    for col in recap_cols.values():
        for row in range(student_start_row, student_end_row + 1):
            cell = ws.cell(row=row, column=col)
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None

    # Clear summary and recap rows if present
    total_col = date_start_col + DATE_COLUMNS
    jumlah_row = _find_row_by_label(ws, "JUMLAH SISWA HADIR", start_row=1)
    if jumlah_row:
        for col in range(date_start_col, total_col + 1):
            cell = ws.cell(row=jumlah_row, column=col)
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None
    recap_title_row = _find_row_by_label(ws, "REKAPITULASI KETIDAKHADIRAN TIAP HARI", start_row=1)
    recap_start = recap_title_row + 1 if recap_title_row else 1
    for label in ("SAKIT", "IZIN", "ALFA"):
        row_num = _find_row_by_label(ws, label, start_row=recap_start)
        if not row_num:
            continue
        for col in range(date_start_col, total_col + 1):
            cell = ws.cell(row=row_num, column=col)
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None


def build_semester_workbook(
    template_path: Path,
    *,
    months: Iterable[Tuple[int, int]],
    students: List[Dict[str, Any]],
    attendance_data: Dict[Tuple[int, int], Dict[Tuple[int, int], str]],
    school_name: Optional[str],
    academic_year: Optional[str],
    class_label: Optional[str],
    teacher_name: Optional[str],
    teacher_nip: Optional[str],
    headmaster_name: Optional[str],
    headmaster_nip: Optional[str],
) -> "Workbook":
    if load_workbook is None:  # pragma: no cover - handled by caller
        raise RuntimeError("openpyxl belum terpasang.")

    wb = load_workbook(template_path)
    template_ws = wb.active

    header_row, date_row, student_start_row, student_end_row = _apply_static_context(
        template_ws,
        students=students,
        school_name=school_name,
        academic_year=academic_year,
        class_label=class_label,
        teacher_name=teacher_name,
        teacher_nip=teacher_nip,
        headmaster_name=headmaster_name,
        headmaster_nip=headmaster_nip,
    )
    date_start_col = _find_date_start_col(template_ws, header_row)
    bulan_cell = _find_bulan_cell(template_ws)
    template_month_year = _parse_bulan_label(bulan_cell.value if bulan_cell else None)
    date_weekday_style, date_weekend_style = _extract_weekday_weekend_styles(
        template_ws,
        wb,
        date_row=date_row,
        date_start_col=date_start_col,
        reference_row=date_row,
        template_month_year=template_month_year,
    )
    student_weekday_style, student_weekend_style = _extract_weekday_weekend_styles(
        template_ws,
        wb,
        date_row=date_row,
        date_start_col=date_start_col,
        reference_row=student_start_row,
        template_month_year=template_month_year,
    )

    month_list = list(months)
    for idx, (year, month) in enumerate(month_list):
        if idx == 0:
            ws = template_ws
        else:
            ws = wb.copy_worksheet(template_ws)
        month_label = INDONESIAN_MONTH_NAMES.get(month, str(month))
        ws.title = f"{month_label} {year}"
        _apply_month_context(ws, year=year, month=month, header_row=header_row, date_row=date_row)
        recap_cols = _find_recap_columns(ws, date_row)
        _clear_attendance_grid(
            ws,
            student_start_row=student_start_row,
            student_end_row=student_end_row,
            date_start_col=date_start_col,
            recap_cols=recap_cols,
        )
        _apply_weekend_styles(
            ws,
            year=year,
            month=month,
            date_row=date_row,
            date_start_col=date_start_col,
            student_start_row=student_start_row,
            student_end_row=student_end_row,
            date_weekday_style=date_weekday_style,
            date_weekend_style=date_weekend_style,
            student_weekday_style=student_weekday_style,
            student_weekend_style=student_weekend_style,
        )
        month_entries = attendance_data.get((year, month), {})
        if month_entries:
            _apply_attendance_data(
                ws,
                students=students,
                month_entries=month_entries,
                header_row=header_row,
                student_start_row=student_start_row,
                date_start_col=date_start_col,
                year=year,
                month=month,
            )

    return wb


def generate_semester_excel(
    template_path: Path,
    *,
    months: Iterable[Tuple[int, int]],
    students: List[Dict[str, Any]],
    attendance_data: Dict[Tuple[int, int], Dict[Tuple[int, int], str]],
    school_name: Optional[str],
    academic_year: Optional[str],
    class_label: Optional[str],
    teacher_name: Optional[str],
    teacher_nip: Optional[str],
    headmaster_name: Optional[str],
    headmaster_nip: Optional[str],
) -> BytesIO:
    wb = build_semester_workbook(
        template_path,
        months=months,
        students=students,
        attendance_data=attendance_data,
        school_name=school_name,
        academic_year=academic_year,
        class_label=class_label,
        teacher_name=teacher_name,
        teacher_nip=teacher_nip,
        headmaster_name=headmaster_name,
        headmaster_nip=headmaster_nip,
    )
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
